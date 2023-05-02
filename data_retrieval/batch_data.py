import psycopg2
import io
from PIL import Image
import openpyxl

conn = psycopg2.connect(host = "dpg-ch7rfo82qv2864obrud0-a.oregon-postgres.render.com",
                        port = 5432, 
                        database = "image_database", 
                        user = "image_database_user", 
                        password = "2ygmOfAGddzefeniRaRWRtR77mQwyvUu")

query = "SELECT user_name, file_name, img FROM image_database"

cur = conn.cursor()
cur.execute(query)
data = cur.fetchall()

wb = openpyxl.Workbook()
ws = wb.active

for row_num, (user_name, file_name, img) in enumerate(data):
    image = Image.open(io.BytesIO(img))
    image.save(f"{file_name}")
    
    ws.cell(row = row_num + 1, column = 1, value = user_name)
    ws.cell(row = row_num + 1, column = 2, value = file_name)

wb.save("data.xlsx")

cur.close()
conn.close()